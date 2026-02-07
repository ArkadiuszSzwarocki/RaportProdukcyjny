"""
Serwis do generowania raportów (PDF, Excel, txt) dla zmian produkcyjnych
"""
import json
from datetime import datetime
from io import BytesIO, StringIO
import csv

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class RaportService:
    """Serwis do generowania raportów zmian produkcyjnych"""
    
    @staticmethod
    def generate_email_text(zmiana_data: dict) -> str:
        """Generuj tekst do wklejenia w email"""
        lines = []
        lines.append("=" * 80)
        lines.append("RAPORT Z ZMIANY - PRODUKCJA")
        lines.append("=" * 80)
        lines.append("")
        
        lines.append(f"DATA ZMIANY: {zmiana_data.get('data', 'N/A')}")
        lines.append(f"SEKCJA: {zmiana_data.get('sekcja', 'N/A')}")
        lines.append(f"LIDER: {zmiana_data.get('lider_name', 'N/A')}")
        lines.append("")
        
        # Obsada per sekcja
        wszystkie_obsady = zmiana_data.get('wszystkie_obsady', {})
        if wszystkie_obsady:
            lines.append("-" * 80)
            lines.append("OBSADA PRACOWNIKÓW:")
            lines.append("-" * 80)
            for sekcja_name, obsada in wszystkie_obsady.items():
                lines.append(f"\n  {sekcja_name.upper()}:")
                if obsada:
                    for p in obsada:
                        lines.append(f"    • {p.get('imie_nazwisko', 'N/A')} ({p.get('rola', 'N/A')})")
                else:
                    lines.append("    Brak przypisanych pracowników")
            lines.append("")
        else:
            # Fallback do starego formatu dla wstecznej kompatybilności
            lines.append("-" * 80)
            lines.append("PRACOWNICY:")
            lines.append("-" * 80)
            pracownicy = zmiana_data.get('pracownicy', [])
            if pracownicy:
                for p in pracownicy:
                    lines.append(f"  • {p.get('imie', 'N/A')}")
            else:
                lines.append("  Brak danych")
            lines.append("")
        
        # Awarie i usterki
        awarie = zmiana_data.get('awarie', [])
        if awarie:
            lines.append("-" * 80)
            lines.append("AWARIE / USTERKI / NIEOBECNOŚCI:")
            lines.append("-" * 80)
            for a in awarie:
                lines.append(f"\n  [{a.get('sekcja', 'N/A')}] {a.get('typ', 'N/A').upper()}")
                lines.append(f"  Opis: {a.get('opis', 'N/A')}")
                lines.append(f"  Czas: {a.get('data_wpisu', 'N/A')}")
                lines.append(f"  Pracownik: {a.get('pracownik', 'N/A')}")
            lines.append("")
        
        lines.append("-" * 80)
        lines.append("PLANY PRODUKCJI:")
        lines.append("-" * 80)
        plany = zmiana_data.get('plany', [])
        if plany:
            for plan in plany:
                sekcja_info = f" ({plan.get('sekcja', '')})" if plan.get('sekcja') else ""
                lines.append(f"\n  Produkt: {plan.get('produkt', 'N/A')}{sekcja_info}")
                lines.append(f"  Plan: {plan.get('tonaz', 'N/A')} t")
                lines.append(f"  Wykonanie: {plan.get('tonaz_wykonania', 'N/A')} t")
                lines.append(f"  Status: {plan.get('status', 'N/A')}")
                
                palety = plan.get('palety', [])
                if palety:
                    lines.append(f"  Palety ({len(palety)}):")
                    for pal in palety[:10]:  # Limit do 10 palet
                        lines.append(f"    - {pal.get('waga', 'N/A')} kg (dodana: {pal.get('data_dodania', 'N/A')})")
                    if len(palety) > 10:
                        lines.append(f"    ... i {len(palety) - 10} więcej palet")
        else:
            lines.append("  Brak danych")
        lines.append("")
        
        lines.append("-" * 80)
        lines.append("NOTATKI LIDERA:")
        lines.append("-" * 80)
        notatki = zmiana_data.get('notatki', '')
        if notatki:
            lines.append(notatki)
        else:
            lines.append("  Brak notatek")
        lines.append("")
        
        lines.append("=" * 80)
        lines.append(f"Raport wygenerowany: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 80)
        
        return "\n".join(lines)
    
    @staticmethod
    def generate_excel(zmiana_data: dict) -> bytes:
        """Generuj raport Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl nie jest zainstalowana. Zainstaluj: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Zmiana"
        
        # Ustawienia
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws['A1'] = "RAPORT Z ZMIANY - PRODUKCJA"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        row = 3
        ws[f'A{row}'] = "DATA ZMIANY:"
        ws[f'B{row}'] = zmiana_data.get('data', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "SEKCJA:"
        ws[f'B{row}'] = zmiana_data.get('sekcja', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "LIDER:"
        ws[f'B{row}'] = zmiana_data.get('lider_name', 'N/A')
        row += 2
        
        # Obsada per sekcja (nowa sekcja)
        wszystkie_obsady = zmiana_data.get('wszystkie_obsady', {})
        if wszystkie_obsady:
            ws[f'A{row}'] = "OBSADA PRACOWNIKÓW PER SEKCJA"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            for sekcja, obsada in wszystkie_obsady.items():
                ws[f'A{row}'] = f"  {sekcja}:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                if obsada:
                    for p in obsada:
                        ws[f'A{row}'] = f"    • {p.get('imie_nazwisko', 'N/A')} ({p.get('rola', 'N/A')})"
                        row += 1
                else:
                    ws[f'A{row}'] = "    Brak przypisanych pracowników"
                    row += 1
            row += 1
        else:
            # Fallback dla starszego formatu
            ws[f'A{row}'] = "PRACOWNICY"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            pracownicy = zmiana_data.get('pracownicy', [])
            for p in pracownicy:
                ws[f'A{row}'] = p.get('imie', 'N/A')
                row += 1
            row += 1
        
        # Awarie/usterki (nowa sekcja)
        awarie = zmiana_data.get('awarie', [])
        if awarie:
            ws[f'A{row}'] = "AWARIE / USTERKI / NIEOBECNOŚCI"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            headers = ['Sekcja', 'Typ', 'Opis', 'Czas', 'Pracownik']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            for a in awarie:
                ws[f'A{row}'] = a.get('sekcja', '')
                ws[f'B{row}'] = a.get('typ', 'N/A')
                ws[f'C{row}'] = a.get('opis', '')
                ws[f'D{row}'] = a.get('data_wpisu', '')
                ws[f'E{row}'] = a.get('pracownik', 'N/A')
                row += 1
            row += 1
        
        # Plany
        ws[f'A{row}'] = "PLANY PRODUKCJI"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        headers = ['Sekcja', 'Produkt', 'Plan (t)', 'Wykonanie (t)', 'Status', 'Ilość palet']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        plany = zmiana_data.get('plany', [])
        for plan in plany:
            sekcja_val = plan.get('sekcja', '')
            ws[f'A{row}'] = sekcja_val if sekcja_val else ''
            ws[f'B{row}'] = plan.get('produkt', 'N/A')
            ws[f'C{row}'] = plan.get('tonaz', 'N/A')
            ws[f'D{row}'] = plan.get('tonaz_wykonania', 'N/A')
            ws[f'E{row}'] = plan.get('status', 'N/A')
            ws[f'F{row}'] = len(plan.get('palety', []))
            row += 1
        
        # Auto adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # Notatki
        row += 1
        ws[f'A{row}'] = "NOTATKI LIDERA"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        notatki = zmiana_data.get('notatki', '')
        ws[f'A{row}'] = notatki if notatki else "Brak notatek"
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws[f'A{row}'].border = border
        
        # Zapisz do BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def generate_pdf(zmiana_data: dict) -> bytes:
        """Generuj raport PDF"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab nie jest zainstalowana. Zainstaluj: pip install reportlab")
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F497D'),
            spaceAfter=30,
            alignment=1  # center
        )
        
        # Title
        story.append(Paragraph("RAPORT Z ZMIANY - PRODUKCJA", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Basic info
        info_data = [
            ['Data zmiany:', zmiana_data.get('data', 'N/A')],
            ['Sekcja:', zmiana_data.get('sekcja', 'N/A')],
            ['Lider:', zmiana_data.get('lider_name', 'N/A')]
        ]
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Obsada per sekcja (nowa sekcja)
        wszystkie_obsady = zmiana_data.get('wszystkie_obsady', {})
        if wszystkie_obsady:
            story.append(Paragraph("Obsada Pracowników per Sekcja", styles['Heading2']))
            for sekcja, obsada in wszystkie_obsady.items():
                text = f"<b>{sekcja}:</b> "
                if obsada:
                    text += ", ".join([f"{p.get('imie_nazwisko', 'N/A')} ({p.get('rola', 'N/A')})" for p in obsada])
                else:
                    text += "Brak przypisanych pracowników"
                story.append(Paragraph(text, styles['Normal']))
            story.append(Spacer(1, 0.3*cm))
        
        # Awarie/usterki table (nowa sekcja)
        awarie = zmiana_data.get('awarie', [])
        if awarie:
            story.append(Paragraph("Awarie / Usterki / Nieobecności", styles['Heading2']))
            awarie_data = [['Sekcja', 'Typ', 'Opis', 'Czas', 'Pracownik']]
            for a in awarie:
                awarie_data.append([
                    a.get('sekcja', '')[:10],
                    a.get('typ', 'N/A')[:10],
                    a.get('opis', '')[:20],
                    a.get('data_wpisu', '')[:8],
                    a.get('pracownik', 'N/A')[:15]
                ])
            
            awarie_table = Table(awarie_data, colWidths=[2*cm, 2*cm, 4*cm, 2*cm, 3*cm])
            awarie_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CC0000')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFE6E6')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            story.append(awarie_table)
            story.append(Spacer(1, 0.3*cm))
        
        # Plans table
        story.append(Paragraph("Plany Produkcji", styles['Heading2']))
        plany = zmiana_data.get('plany', [])
        if plany:
            plans_data = [['Sekcja', 'Produkt', 'Plan (t)', 'Wykonanie (t)', 'Status', 'Palety']]
            for plan in plany:
                sekcja_val = plan.get('sekcja', '')
                plans_data.append([
                    sekcja_val[:10] if sekcja_val else '',
                    plan.get('produkt', 'N/A')[:15],
                    str(plan.get('tonaz', 'N/A')),
                    str(plan.get('tonaz_wykonania', 'N/A')),
                    plan.get('status', 'N/A'),
                    str(len(plan.get('palety', [])))
                ])
            
            plans_table = Table(plans_data, colWidths=[2.5*cm, 5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.5*cm])
            plans_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            story.append(plans_table)
            story.append(Spacer(1, 0.5*cm))
        
        # Notes
        story.append(Paragraph("Notatki Lidera", styles['Heading2']))
        notatki = zmiana_data.get('notatki', 'Brak notatek')
        story.append(Paragraph(notatki if notatki else "Brak notatek", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output.getvalue()

